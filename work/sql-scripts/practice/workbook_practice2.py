from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = 'Grade_report'

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

wb.save('Grades2.xlsx')