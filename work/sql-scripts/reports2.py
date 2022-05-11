import psycopg2
import uuid
import base58
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import base64
wb = Workbook()
ws = wb.active
ws.title = 'Practice_report'


con = psycopg2.connect(
    host="db1-processing-prod-replica.cirlgfwr1ct4.us-east-1.rds.amazonaws.com",
    database="dashboard",
    user='vbradley', 
    password= 'Van123!braD'
    )

#cursor
cur = con.cursor()

#esecute th query
# cur.execute('SELECT * FROM processing.applications LIMIT 1')
 
cur.execute(''' SELECT
  du.auth0_user_id,
  du.email_address,
  substring(convert_from(decode(m.encrypted_password, 'base64'), 'UTF8') from '(.*):') AS user_id,
  api.env AS environment,
  LEAST(du.created_at, du.v1_created_at) AS user_created_at,
  du.enabled AS user_enabled,
  LEAST(api.created_at, api.v1_created_at) AS api_created_at,
  api.enabled AS api_enabled,
  api.role_name AS api_role
FROM
  dashboard_user AS du
  JOIN dashboard AS d ON du.dashboard_id = d.id
  JOIN membership AS m on du.id = m.dashboard_user_id 
  JOIN api_credential AS api ON m.api_credential_id = api.id
WHERE
  d.name = 'lightspeed'
  ORDER BY api.env; ''')

rows = cur.fetchall()

# for r in rows:
#     print(r)

headings = ['Auth0', 'Email', 'USXXX', 'Env', 'Created At', 'User_Enabled', 'API Created At', 'API Enabled', 'API Role']
ws.append(headings)
ids = []
for r in rows:
    ids.append(r[2])
    new_list = list(r)
    # print(new_list)
    new_list[3] = r[3]
    c_date = str(r[4])
    u_date = str(r[6])
    new_list[4] = c_date[0:10]
    new_list[6] = u_date[0:10]
    if r[3] == None:
        new_list[3] = 'NULL'
    if r[5] == None:
        new_list[5] = 'NULL'
    if r[7] == None:
        new_list[7] = 'NULL'
    if r[8] == None:
        new_list[8] = 'NULL'
    ws.append(new_list)
   
    # print(r[3])


#close cursor
cur.close()

#close the connection
con.close()


wb.save('Practice_report.xlsx')



# wb = load_workbook('Pratice_report.py')
# ws = wb['Sheet2']

# con = psycopg2.connect(
#     host="db1-processing-prod-replica.cirlgfwr1ct4.us-east-1.rds.amazonaws.com",
#     database="authentication",
#     user='vbradley', 
#     password= 'Van123!braD'
#     )
# for id in ids:
#     # print(id)
#     print(base64.b64encode(id))
   







