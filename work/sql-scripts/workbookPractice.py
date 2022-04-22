from openpyxl import Workbook, load_workbook
from openpyxl.utils  import get_column_letter

# wb = load_workbook('Grades.xlsx')
# ws = wb['Sheet2']
# ws['A2'] = 'Test'
# print(ws['A2'].value)
# wb.save('Grades.xlsx')
# print(wb['Sheet2'])

# print(ws)


# wb.create_sheet("Test")
# print(wb.sheetnames)

#Create new workbook

# wb = Workbook()
# ws = wb.active
# ws.title = 'Data'

# #Adding and appending rows
# ws.append(['ID', 'API Key', 'ENV'])
# ws.append(['ID', 'API Key', 'ENV'])
# ws.append(['ID', 'API Key', 'ENV'])
# ws.append(['End'])
# wb.save('tim.xlsx')


#Looping through a workbook
# wb = load_workbook('tim.xlsx')
# ws = wb.active

# for row in range(1,11):
#     for col in range(1,5):
#         char = get_column_letter(col)
#         ws[char + str(row)] = char + str(row)
#         # print(ws[char + str(row)].value)


# wb.save('tim.xlsx')


#Merging cells

wb = load_workbook('tim.xlsx')
ws = wb.active

# ws.merge_cells("A1:D1")

#Un-merge cells
# ws.unmerge_cells("A1:D1")

# ws.insert_rows(7)
# ws.insert_rows(7)

# ws.delete_rows(7)


# ws.insert_cols(2)
ws.delete_cols(2)
wb.save('tim.xlsx')
